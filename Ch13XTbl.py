#Needs work, incomplete!

import openpyxl
from openpyxl.styles import Font
wb =  openpyxl.load_workbook('MultiT.xlsx')
wb.sheetnames
sheet = wb.active

n =int(input())
rc = n
rr = n + 1

def print_rows():
    for row in sheet.iter_rows(min_row=1, max_col=rr, max_row=rr, values_only=True):
        print(row)

def populatecells():
    for i in range(1, rr):
        sheet['A' + str(i+1)] = i
        for j in range(1, rr):
            sheet.cell(row=i+1, column=j+1).value = i*j  
  
populatecells()
print_rows()

sheet['A2'].font=Font(bold=True) #bold, example
wb.save('MultiT.xlsx')

#can't figure out how to number columns (across) 
# from user input so I left it as none's :(
